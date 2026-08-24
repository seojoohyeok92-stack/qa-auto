from __future__ import annotations

import asyncio
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .engine import AnswerEngine
from .config import ROOT
from .config_excel import append_product_code_if_missing
from .learning import (
    REVIEW_PATH,
    append_answer_history,
    append_learning_review,
    build_learning_candidate,
    load_latest_history,
)
from .naver_client import (
    NaverCommerceClient,
    NaverConfig,
    merge_detail,
    normalize_question,
    raw_json,
)
from .kakao_notify import notify_qna_safely
import hashlib
import sqlite3

PROCESS_DB = ROOT / "data" / "naver_qna_processed.sqlite3"

MASTER_REPORT_PATH = (
    ROOT
    / "outputs"
    / "naver_api"
    / "naver_qna_history.xlsx"
)


def build_question_key(question) -> str:
    question_id = str(
        question.question_id or ""
    ).strip()

    if question_id:
        return f"naver:{question_id}"

    raw = "\n".join(
        [
            str(question.product or "").strip(),
            str(question.option_name or "").strip(),
            str(question.question or "").strip(),
        ]
    )

    digest = hashlib.sha256(
        raw.encode("utf-8")
    ).hexdigest()

    return f"hash:{digest}"


def connect_process_db() -> sqlite3.Connection:
    PROCESS_DB.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    conn = sqlite3.connect(
        PROCESS_DB,
        timeout=30,
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS processed_questions (
            question_key TEXT PRIMARY KEY,
            question_id TEXT,
            action TEXT NOT NULL,
            product TEXT,
            question TEXT,
            processed_at TEXT NOT NULL
        )
        """
    )

    conn.commit()
    return conn


def is_question_processed(
    question_key: str,
) -> bool:
    conn = connect_process_db()

    try:
        row = conn.execute(
            """
            SELECT 1
            FROM processed_questions
            WHERE question_key = ?
            LIMIT 1
            """,
            (question_key,),
        ).fetchone()

        return row is not None

    finally:
        conn.close()



def mark_question_processed(
    *,
    question_key: str,
    question_id: str,
    action: str,
    product: str,
    question: str,
) -> None:
    conn = connect_process_db()

    try:
        conn.execute(
            """
            INSERT OR REPLACE INTO processed_questions (
                question_key,
                question_id,
                action,
                product,
                question,
                processed_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                question_key,
                question_id,
                action,
                product,
                question,
                datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
            ),
        )

        conn.commit()

    finally:
        conn.close()

@dataclass
class NaverRunOptions:
    limit: int = 20
    page_size: int = 20
    days: int = 7
    post: bool = False
    fetch_detail: bool = True
    output: Path | None = None


async def run_naver_qna(options: NaverRunOptions) -> Path:
    client = NaverCommerceClient(NaverConfig.from_env(dry_run=not options.post, days=options.days))
    engine = AnswerEngine()

    records: list[dict[str, Any]] = []
    seen_question_keys: set[str] = set()
    page = 1
    while len(records) < options.limit:
        raw_items = await client.fetch_unanswered_questions(page=page, size=options.page_size, days=options.days)
        if not raw_items:
            break
        for raw in raw_items:
            if len(records) >= options.limit:
                break
            summary_question = normalize_question(raw)
            detail = None
            if options.fetch_detail and summary_question.question_id:
                detail = await client.fetch_question_detail(summary_question.question_id)
            merged = merge_detail(raw, detail)
            question = normalize_question(merged)

            question_key = build_question_key(
                question
            )

            if question_key in seen_question_keys:
                print(
                    "[SKIP] 같은 실행 내 중복 문의: "
                    f"{question.question_id or question_key}"
                )
                continue

            if is_question_processed(
                question_key
            ):
                print(
                    "[SKIP] 이미 처리한 네이버 문의: "
                    f"{question.question_id or question_key}"
                )
                continue

            seen_question_keys.add(question_key)

            if (
                not question.question_id
                or not question.product
                or not question.question
            ):
                record = build_record(
                    question,
                    None,
                    "skip_missing_fields",
                    False,
                )

                record["question_key"] = question_key
                append_product_code_if_missing(ROOT, record)
                records.append(record)

                notify_qna_safely(
                    title="[네이버 Q&A 필수정보 누락]",
                    product=question.product or "",
                    option_name=question.option_name or "",
                    question=question.question or "",
                    answer="",
                    reason=(
                        "네이버 응답에서 question_id, 상품명 또는 "
                        "문의 내용을 확인하지 못했습니다."
                    ),
                    action="skip_missing_fields",
                    notify_key=question_key,
                )

                continue

            result = engine.answer(
                question.product,
                question.question,
                question.option_name,
            )

            posted = False
            action = "no_answer"

            if result.status == "답변 가능" and result.answer:
                posted = await client.post_answer(
                    question.question_id,
                    result.answer,
                )

                action = (
                    "posted"
                    if posted
                    else ("dry_run" if not options.post else "post_failed")
                )

            record = build_record(
                question,
                result,
                action,
                posted,
            )
            record["question_key"] = question_key
            append_product_code_if_missing(ROOT, record)
            records.append(record)

            # 카카오 알림 제목 결정
            if action == "posted":
                kakao_title = "[네이버 Q&A 답변 등록 완료]"

            elif action == "dry_run":
                kakao_title = "[네이버 Q&A 신규 문의]"

            elif action == "post_failed":
                kakao_title = "[네이버 Q&A 답변 등록 실패]"

            else:
                kakao_title = "[네이버 Q&A 자동답변 보류]"

            # 디스패처 대기열에 카카오 알림 등록
            notify_qna_safely(
                title=kakao_title,
                product=question.product,
                option_name=question.option_name,
                question=question.question,
                answer=result.answer or "",
                reason=result.reason or "",
                action=action,
                inquiry_id=str(
                    question.question_id or ""
                ),
                notify_key=question_key,
            )

        page += 1

    output = (
        options.output
        or default_output_path()
    )

    save_naver_report(
        records,
        output,
    )

    # 엑셀 저장 성공 후 처리 완료 기록
    for record in records:
        action = str(
            record.get("action") or ""
        )

        if action not in {
            "posted",
            "no_answer",
            "skip_missing_fields",
        }:
            continue

        mark_question_processed(
            question_key=str(
                record.get("question_key") or ""
            ),
            question_id=str(
                record.get("question_id") or ""
            ),
            action=action,
            product=str(
                record.get("product") or ""
            ),
            question=str(
                record.get("question") or ""
            ),
        )

    if records:
        append_answer_history(records)

    return output


def run_naver_qna_sync(options: NaverRunOptions) -> Path:
    return asyncio.run(run_naver_qna(options))


async def collect_learning_review(options: NaverRunOptions) -> Path:
    client = NaverCommerceClient(NaverConfig.from_env(dry_run=True, days=options.days))
    history = load_latest_history()
    candidates: list[dict[str, Any]] = []

    page = 1
    while len(candidates) < options.limit:
        raw_items = await client.fetch_questions(page=page, size=options.page_size, days=options.days, answered=True)
        if not raw_items:
            break
        for raw in raw_items:
            if len(candidates) >= options.limit:
                break
            summary_question = normalize_question(raw)
            detail = None
            if options.fetch_detail and summary_question.question_id:
                detail = await client.fetch_question_detail(summary_question.question_id)
            merged = merge_detail(raw, detail)
            question = normalize_question(merged)
            candidate = build_learning_candidate(question, history)
            if candidate:
                candidates.append(candidate)
        page += 1

    return append_learning_review(candidates, options.output or REVIEW_PATH)


def collect_learning_review_sync(options: NaverRunOptions) -> Path:
    return asyncio.run(collect_learning_review(options))


def build_record(question, result, action: str, posted: bool) -> dict[str, Any]:
    processed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {
        "processed_at": processed_at,
        "question_id": question.question_id,
        "action": action,
        "posted": "Y" if posted else "N",
        "product": question.product,
        "option_name": question.option_name,
        "question": question.question,
        "program_status": result.status if result else "스킵",
        "program_answer": result.answer if result else "",
        "program_reason": result.reason if result else "네이버 응답에서 필수값을 확인하지 못했습니다.",
        "category": result.category if result else "필수값누락",
        "question_count": result.question_count if result else 0,
        "question_breakdown": result.question_breakdown if result else "",
        "provider": result.provider if result else "system",
        "order_id": question.order_id,
        "product_order_id": question.product_order_id,
        "channel_product_no": question.channel_product_no,
        "origin_product_no": question.origin_product_no,
        "raw_json": raw_json(question.raw),
    }


def default_output_path() -> Path:
    return MASTER_REPORT_PATH


def save_naver_report(
    records: list[dict[str, Any]],
    output: str | Path,
) -> Path:
    output = Path(output)
    output.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    headers = [
        "처리시각",
        "question_id",
        "action",
        "posted",
        "상품",
        "옵션",
        "문의",
        "답변여부",
        "답변",
        "판단사유",
        "질문유형",
        "질문수",
        "개별질문요약",
        "판단방식",
        "order_id",
        "product_order_id",
        "channel_product_no",
        "origin_product_no",
        "raw_json",
    ]

    key_map = {
        "처리시각": "processed_at",
        "상품": "product",
        "옵션": "option_name",
        "문의": "question",
        "답변여부": "program_status",
        "답변": "program_answer",
        "판단사유": "program_reason",
        "질문유형": "category",
        "질문수": "question_count",
        "개별질문요약": "question_breakdown",
        "판단방식": "provider",
    }

    # 새로 처리한 문의가 없으면 기존 파일 그대로 반환
    if not records:
        print(
            "[EXCEL] 새로 기록할 문의가 없습니다."
        )

        if not output.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "naver_qna_result"
            ws.append(headers)
            style_sheet(ws)
            wb.save(output)

        return output

    if output.exists():
        wb = load_workbook(output)

        if "naver_qna_result" in wb.sheetnames:
            ws = wb["naver_qna_result"]
        else:
            ws = wb.create_sheet(
                "naver_qna_result"
            )
            ws.append(headers)

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "naver_qna_result"
        ws.append(headers)

    for record in records:
        row = [
            record.get(
                key_map.get(header, header),
                "",
            )
            for header in headers
        ]

        ws.append(row)

    style_sheet(ws)

    # 가장 최근 처리 건이 위쪽에 보이게 하려면
    # 현재 방식 대신 정렬 로직을 별도로 넣을 수 있음
    wb.save(output)

    print(
        f"[EXCEL] 신규 문의 {len(records)}건 누적 저장"
    )
    print(
        f"[EXCEL] 파일: {output}"
    )

    return output


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="244C5A")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [18, 22, 14, 8, 44, 28, 64, 14, 86, 48, 24, 8, 60, 14, 20, 22, 22, 22, 90]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
