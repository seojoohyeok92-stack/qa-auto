from __future__ import annotations

import argparse
import hashlib
import json
import sys
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.list_answer_templates import RuleVisitor, catalog
import ast


DEFAULT_ROOTS = (PROJECT_ROOT,)
DEFAULT_ZIP = PROJECT_ROOT / "legacy_answer_source.zip"


def _hash(value: str | bytes) -> str:
    payload = value.encode("utf-8") if isinstance(value, str) else value
    return hashlib.sha256(payload).hexdigest()


def _engine_records(path: Path) -> list[dict[str, Any]]:
    visitor = RuleVisitor()
    visitor.visit(ast.parse(path.read_text(encoding="utf-8")))
    return [
        {
            "template_id": row["template_id"],
            "content_hash": _hash(str(row["answer_text"])),
            "source": str(path),
            "kind": "PYTHON_RULE",
        }
        for row in visitor.rows
    ]


def _json_records(path: Path) -> list[dict[str, Any]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    rows: list[dict[str, Any]] = []

    def walk(value: Any, keys: list[str]) -> None:
        if isinstance(value, dict):
            for key, child in value.items():
                walk(child, [*keys, str(key)])
        elif isinstance(value, list):
            for index, child in enumerate(value):
                walk(child, [*keys, str(index)])
        elif isinstance(value, str) and keys and (
            keys[-1].endswith("answer")
            or keys[-1] in {"신규주문안내", "기존주문안내", "답변본문"}
        ):
            rows.append(
                {
                    "template_id": f"json:{path.name}:{'.'.join(keys)}",
                    "content_hash": _hash(value.strip()),
                    "source": str(path),
                    "kind": "JSON_CONFIG",
                }
            )

    walk(data, [])
    return rows


def _xlsx_records(path: Path) -> list[dict[str, Any]]:
    if path.name.lower() != "configuration.xlsx":
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for sheet_name in ("학습답변룰", "설치배송일정"):
            if sheet_name not in workbook.sheetnames:
                continue
            values = list(workbook[sheet_name].iter_rows(values_only=True))
            if not values:
                continue
            headers = [str(value or "") for value in values[0]]
            for index, value_row in enumerate(values[1:], start=2):
                item = dict(zip(headers, value_row))
                if not any(value not in (None, "") for value in value_row):
                    continue
                answer_parts = [
                    str(item.get(key) or "").strip()
                    for key in ("답변본문", "신규주문안내", "기존주문안내")
                    if str(item.get(key) or "").strip()
                ]
                for part_index, answer in enumerate(answer_parts, start=1):
                    rows.append(
                        {
                            "template_id": (
                                f"xlsx:{sheet_name}:{index}:{part_index}"
                            ),
                            "content_hash": _hash(answer),
                            "source": str(path),
                            "kind": "XLSX_RULE",
                        }
                    )
    finally:
        workbook.close()
    return rows


def _directory_records(root: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    engine_paths = [
        root / "answer" / "engine.py",
        root / "qna_auto" / "engine.py",
    ]
    for path in engine_paths:
        if path.is_file():
            rows.extend(_engine_records(path))
    for path in root.rglob("*.json"):
        if path.name.startswith(".env") or any(
            part.lower() in {"logs", "data", "outputs"}
            for part in path.parts
        ):
            continue
        rows.extend(_json_records(path))
    for path in root.rglob("configuration.xlsx"):
        rows.extend(_xlsx_records(path))
    return rows


def _zip_summary(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return {"path": str(path), "exists": False}
    entries: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as archive:
        for info in archive.infolist():
            normalized = info.filename.replace("\\", "/")
            if normalized.endswith(
                ("qna_auto/engine.py", "qna_auto_configs/configuration.xlsx")
            ):
                entries.append(
                    {
                        "entry": normalized,
                        "sha256": _hash(archive.read(info)),
                        "size": info.file_size,
                    }
                )
    return {"path": str(path), "exists": True, "key_entries": entries}


def audit() -> dict[str, Any]:
    projects: list[dict[str, Any]] = []
    all_records: list[dict[str, Any]] = []
    for root in DEFAULT_ROOTS:
        if not root.is_dir():
            continue
        rows = _directory_records(root)
        all_records.extend({**row, "project": str(root)} for row in rows)
        projects.append(
            {
                "path": str(root),
                "template_records": len(rows),
                "unique_content": len({row["content_hash"] for row in rows}),
                "kinds": {
                    kind: sum(row["kind"] == kind for row in rows)
                    for kind in sorted({row["kind"] for row in rows})
                },
            }
        )
    current_hashes = {
        _hash(str(row.get("answer_text") or "").strip())
        for row in catalog()
        if str(row.get("answer_text") or "").strip()
    }
    legacy_records = [
        row
        for row in all_records
        if not row["project"].endswith("Q&A auto")
    ]
    legacy_hashes = {row["content_hash"] for row in legacy_records}
    return {
        "projects": projects,
        "zip_source": _zip_summary(DEFAULT_ZIP),
        "records_before_deduplication": len(all_records),
        "records_after_content_deduplication": len(
            {row["content_hash"] for row in all_records}
        ),
        "legacy_unique_content": len(legacy_hashes),
        "legacy_content_already_in_current": len(
            legacy_hashes & current_hashes
        ),
        "legacy_content_missing_from_current": len(
            legacy_hashes - current_hashes
        ),
        "legacy_missing_hashes": sorted(legacy_hashes - current_hashes),
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Read-only audit of legacy answer template sources."
    )
    parser.parse_args()
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    print(json.dumps(audit(), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
