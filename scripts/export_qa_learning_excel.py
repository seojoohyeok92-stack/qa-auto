r"""Export the operational Q&A and Learning data to one reviewable workbook.

Read-only by construction: both databases are opened with SQLite's ``mode=ro``
URI *and* ``PRAGMA query_only=ON``, only SELECT statements are issued, and the
file modification times are compared before and after the run.

Nothing here decides anything. Every verdict shown -- staff review, hold
reason, registration state, whether a Learning row may serve as evidence -- is
read from the same production components the pipeline itself uses, so the
workbook cannot disagree with the dashboard:

    ui.answer_status_presenter.build_answer_status   staff review / registration
    answer.hold_reasons.primary_reason               why it was held
    services.learning_privacy_service                the masking policy
    services.learning_evidence_policy                hedged / redaction-token
    services.learning_validity_service               temporary validity status

Usage (PowerShell, from the project root)::

    python .\scripts\export_qa_learning_excel.py
    python .\scripts\export_qa_learning_excel.py --out-dir exports --limit 500
"""
from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from answer.hold_reasons import primary_reason  # noqa: E402
from repositories.database import get_database_path  # noqa: E402
from repositories.product_fact_repository import get_product_facts_path  # noqa: E402
from services.auto_post_pipeline_service import AutoPostPipelineService  # noqa: E402
from services.auto_processing_eligibility_service import (  # noqa: E402
    AutoProcessingEligibilityService,
)
from services.learning_evidence_policy import (  # noqa: E402
    contamination_reason,
    hedge_reason,
    is_hedged,
)
from services.learning_privacy_service import LearningPrivacyService  # noqa: E402
from services.learning_validity_service import (  # noqa: E402
    validity_status,
    validity_summary,
)
from ui.answer_status_presenter import build_answer_status  # noqa: E402


PRIVACY = LearningPrivacyService()

# Columns Excel must not reformat. An order number is an identifier that
# happens to be made of digits: left alone, Excel renders 2026082351391541 as
# 2.02608E+15 and the operator cannot search for it.
TEXT_COLUMNS = frozenset({
    "문의 ID", "상품 ID", "주문번호", "상품주문번호", "Learning ID",
    "모델 코드", "문의 로컬 ID", "연결 문의 ID",
})

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=12)
THIN = Side(style="thin", color="D9D9D9")
CELL_BORDER = Border(bottom=THIN)


# --------------------------------------------------------------------------
# read-only access
# --------------------------------------------------------------------------


def open_readonly(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(
        path.as_uri() + "?mode=ro", uri=True, isolation_level=None, timeout=15
    )
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA query_only = ON")
    if int(connection.execute("PRAGMA query_only").fetchone()[0]) != 1:
        connection.close()
        raise RuntimeError("READ_ONLY_GUARD_FAILED: PRAGMA query_only is not ON")
    return connection


def fingerprint(path: Path) -> tuple[int, int] | None:
    """Size and mtime, so the caller can prove the file was not written to."""

    if not path.exists():
        return None
    stat = path.stat()
    return stat.st_size, stat.st_mtime_ns


def as_json(value: object) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(value or "{}")
    except (TypeError, ValueError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def mask(value: object) -> str:
    """The pipeline's own masking, applied again on the way out.

    Most stored text is already masked at write time; re-running the same
    policy costs nothing and means an Export can never be the thing that
    publishes a customer's phone number. Approved contact numbers
    (1588-3366, 02-706-2678) stay readable because that is what the policy
    itself says -- no separate rule is invented here.
    """

    text = str(value or "").strip()
    return PRIVACY.mask(text) if text else ""


def yes_no(value: object) -> str:
    if value is None or value == "":
        return ""
    return "예" if bool(value) else "아니오"


def korean_datetime(value: object) -> str:
    """One display format for every timestamp column."""

    text = str(value or "").strip()
    if not text:
        return ""
    for pattern in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z",
                    "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ",
                    "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            moment = datetime.strptime(text, pattern)
        except ValueError:
            continue
        if moment.tzinfo is not None:
            moment = moment.astimezone(timezone(KST_OFFSET)).replace(tzinfo=None)
        return moment.strftime("%Y-%m-%d %H:%M")
    return text


from datetime import timedelta  # noqa: E402

KST_OFFSET = timedelta(hours=9)


# --------------------------------------------------------------------------
# Q&A rows
# --------------------------------------------------------------------------


QA_HEADERS = [
    "문의 ID", "문의 로컬 ID", "문의 등록일", "스토어", "문의 유형",
    "상품명", "상품 ID", "모델 코드", "문의 내용", "주문번호", "상품주문번호",
    "문의 분류", "세부 분류", "Intent", "답변 전략",
    "주문조회 필요", "DPS 조회 필요", "주문번호 상태", "주문조회 결과", "DPS 결과",
    "답변 출처", "사용 Rule/Template", "Product Fact 사용", "Product Fact 출처",
    "Learning 사용", "사용 Learning ID",
    "프로그램 생성 답변", "직원 수정 답변", "최종 답변", "실제 등록 답변",
    "Validator", "자동등록 판정", "차단 사유 코드", "보류 사유",
    "직원검토 필요", "자동등록 여부", "등록 상태", "등록 시각",
    "승인 상태", "승인자", "승인 시각", "데이터 품질 경고",
]


def load_qa_rows(
    connection: sqlite3.Connection, *, limit: int | None
) -> list[dict[str, Any]]:
    clause = f"LIMIT {int(limit)}" if limit else ""
    inquiries = connection.execute(
        f"""
        SELECT * FROM inquiries
        ORDER BY COALESCE(registered_at, created_at) DESC, id DESC
        {clause}
        """
    ).fetchall()

    drafts: dict[int, sqlite3.Row] = {}
    for row in connection.execute(
        """
        SELECT * FROM answer_drafts
        ORDER BY inquiry_id, is_active DESC, created_at DESC, id DESC
        """
    ):
        drafts.setdefault(int(row["inquiry_id"]), row)

    posted: dict[int, str] = {}
    for row in connection.execute(
        """
        SELECT inquiry_id, answer_body FROM naver_posted_answers
        WHERE COALESCE(is_current, 1)=1
        ORDER BY inquiry_id, COALESCE(posted_at, last_observed_at) DESC, id DESC
        """
    ):
        posted.setdefault(int(row["inquiry_id"]), str(row["answer_body"] or ""))

    used_learning: dict[int, list[str]] = {}
    for row in connection.execute(
        """
        SELECT inquiry_id, learning_example_id, historical_case_id, reference_kind
        FROM answer_learning_provenance
        WHERE usage_status='USED'
        """
    ):
        label = (
            f"L{row['learning_example_id']}"
            if row["learning_example_id"] is not None
            else f"H{row['historical_case_id']}"
        )
        used_learning.setdefault(int(row["inquiry_id"]), []).append(label)

    dps: dict[int, sqlite3.Row] = {}
    for row in connection.execute(
        """
        SELECT inquiry_id, lookup_status, installation_date, error_code
        FROM dps_lookup_results ORDER BY inquiry_id, id DESC
        """
    ):
        dps.setdefault(int(row["inquiry_id"]), row)

    rows: list[dict[str, Any]] = []
    for raw in inquiries:
        inquiry = dict(raw)
        inquiry["raw_json"] = as_json(inquiry.get("raw_json"))
        local_id = int(inquiry["id"])
        draft_row = drafts.get(local_id)
        draft = dict(draft_row) if draft_row is not None else None
        if draft is not None:
            for key in ("metadata_json", "inquiry_analysis_json",
                        "selected_facts_json", "validator_result_json"):
                draft[key] = as_json(draft.get(key))

        metadata = (draft or {}).get("metadata_json") or {}
        plan = metadata.get("processing_plan") if isinstance(metadata, dict) else {}
        plan = plan if isinstance(plan, dict) else {}
        analysis = plan.get("analysis") if isinstance(plan, dict) else {}
        analysis = analysis if isinstance(analysis, dict) else {}
        guard = metadata.get("product_fact_guard") if isinstance(metadata, dict) else {}
        guard = guard if isinstance(guard, dict) else {}

        route = AutoPostPipelineService._route(draft) if draft else ""
        status = build_answer_status(inquiry=inquiry, draft=draft, route=route)
        verdict = None
        if draft is not None:
            verdict = AutoProcessingEligibilityService().evaluate(
                inquiry=inquiry, draft=draft, route=route
            )
        reason_codes = list(verdict.reasons) if verdict else []
        hold = (
            primary_reason(verdict.reasons, verdict.soft_reasons)
            if verdict and not verdict.safe
            else ""
        )

        dps_row = dps.get(local_id)
        warnings = qa_quality_warnings(inquiry, draft, posted.get(local_id, ""))

        rows.append({
            "문의 ID": str(inquiry.get("external_inquiry_id")
                          or inquiry.get("source_question_id") or ""),
            "문의 로컬 ID": str(local_id),
            "문의 등록일": korean_datetime(inquiry.get("registered_at")),
            "스토어": inquiry.get("store_code") or "",
            "문의 유형": inquiry.get("source_type") or "",
            "상품명": inquiry.get("product_name") or "",
            "상품 ID": str(inquiry.get("product_id")
                         or (inquiry["raw_json"].get("productId") or "")),
            "모델 코드": str(analysis.get("model_code") or ""),
            "문의 내용": mask(inquiry.get("content")),
            "주문번호": str(inquiry.get("order_id") or ""),
            "상품주문번호": str(inquiry.get("product_order_id") or ""),
            "문의 분류": analysis.get("question_category") or "",
            "세부 분류": analysis.get("inquiry_subtype") or "",
            "Intent": analysis.get("detected_intent") or "",
            "답변 전략": analysis.get("answer_strategy") or "",
            "주문조회 필요": yes_no(analysis.get("requires_order_lookup")),
            "DPS 조회 필요": yes_no(analysis.get("requires_dps_lookup")),
            "주문번호 상태": plan.get("order_id_status") or "",
            "주문조회 결과": plan.get("order_lookup_status") or "",
            "DPS 결과": (
                dps_row["lookup_status"] if dps_row is not None
                else plan.get("dps_lookup_status") or ""
            ),
            "답변 출처": (
                metadata.get("generation_mode")
                or (draft or {}).get("provider") or ""
            ),
            "사용 Rule/Template": metadata.get("template_id") or route or "",
            "Product Fact 사용": yes_no(guard.get("current_fact_verified")),
            "Product Fact 출처": guard.get("current_fact_source") or "",
            "Learning 사용": yes_no(bool(used_learning.get(local_id))),
            "사용 Learning ID": ", ".join(used_learning.get(local_id, [])),
            "프로그램 생성 답변": mask((draft or {}).get("original_answer")),
            "직원 수정 답변": mask((draft or {}).get("edited_answer")),
            "최종 답변": mask((draft or {}).get("final_answer")),
            "실제 등록 답변": mask(posted.get(local_id)),
            "Validator": status.validation_label,
            "자동등록 판정": status.registration_label,
            "차단 사유 코드": ", ".join(reason_codes),
            "보류 사유": hold,
            "직원검토 필요": yes_no(status.staff_review_required),
            "자동등록 여부": yes_no(
                str(inquiry.get("post_actor") or "") == "SYSTEM_AUTO_POST"
            ),
            "등록 상태": inquiry.get("post_status") or "",
            "등록 시각": korean_datetime(inquiry.get("posted_at")),
            "승인 상태": inquiry.get("approval_status") or "",
            "승인자": mask(inquiry.get("approved_by")),
            "승인 시각": korean_datetime(inquiry.get("approved_at")),
            "데이터 품질 경고": ", ".join(warnings),
            "_staff_review": bool(status.staff_review_required),
        })
    return rows


def qa_quality_warnings(
    inquiry: dict[str, Any],
    draft: dict[str, Any] | None,
    posted_answer: str = "",
) -> list[str]:
    """Only findings that are objectively checkable in the stored row."""

    found: list[str] = []
    # Deliberately checked against the *stored* text, before this export's
    # own masking runs. A "<masked-phone>" visible in a cell can have two
    # very different causes: the value was already stored that way (a data
    # problem to clean up), or this export just masked a live phone number
    # (privacy working as intended). Only the first is a finding, so the
    # posted answer -- which is stored unmasked and masked on the way out --
    # is checked here in its raw form.
    texts = [
        (draft or {}).get("original_answer"),
        (draft or {}).get("final_answer"),
        posted_answer,
    ]
    if any(contamination_reason(text) for text in texts if text):
        found.append("INTERNAL_PLACEHOLDER")
    if draft is None:
        found.append("NO_DRAFT")
    if not str(inquiry.get("product_id") or "").strip():
        found.append("MISSING_PRODUCT_ID")
    return found


# --------------------------------------------------------------------------
# Learning rows
# --------------------------------------------------------------------------


LEARNING_HEADERS = [
    "Learning ID", "상태", "연결 문의 ID", "질문", "학습 답변",
    "상품명", "상품 ID", "모델 코드", "적용 범위",
    "Positive/Negative", "문체 전용(style_only)", "직원 검증(human_verified)",
    "직원 승인", "직원 수정", "사실 근거 사용 가능",
    "출처(learning_source)", "근거 등급(facts_authority)", "답변 출처(provenance)",
    "평점", "품질 점수",
    "기간성", "이벤트명", "유효 시작", "유효 종료", "유효 상태", "운영 메모",
    "사용 횟수", "마지막 사용", "생성일", "수정일",
    "데이터 품질 경고",
]


def load_learning_rows(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    joined = connection.execute(
        """
        SELECT le.*,
               i.product_id  AS inquiry_product_id,
               i.source_question_id AS inquiry_external_id
        FROM learning_examples le
        LEFT JOIN inquiries i ON i.id = le.inquiry_id
        ORDER BY le.created_at DESC, le.id DESC
        """
    ).fetchall()

    rows: list[dict[str, Any]] = []
    for raw in joined:
        item = dict(raw)
        metadata = as_json(item.get("metadata_json"))
        item["metadata_json"] = metadata
        identity = metadata.get("product_identity")
        identity = identity if isinstance(identity, dict) else {}

        source = str(item.get("learning_source") or "")
        human_verified = metadata.get("human_verified") is True
        approved = human_verified or item.get("approval_history_id") is not None
        edited = source == "APPROVED_EDITED" or bool(
            str(item.get("edited_answer") or "").strip()
        )
        signal = str(metadata.get("learning_signal_type") or "POSITIVE")
        answer = str(item.get("final_answer") or "")
        style_only = bool(item.get("style_only"))
        contaminated = contamination_reason(answer)
        hedged = is_hedged(answer)
        product_id = str(
            identity.get("product_id") or item.get("inquiry_product_id") or ""
        )
        model_code = str(
            item.get("model_code") or identity.get("model_code") or ""
        )

        fact_eligible = bool(
            item.get("active")
            and not style_only
            and human_verified
            and not contaminated
            and not hedged
            and product_id
        )

        warnings: list[str] = []
        if contaminated:
            warnings.append("INTERNAL_PLACEHOLDER")
        if style_only:
            warnings.append("STYLE_ONLY")
        if not model_code:
            warnings.append("MISSING_MODEL_CODE")
        if not product_id:
            warnings.append("MISSING_PRODUCT_IDENTITY")
        if hedged:
            warnings.append("HEDGED_ANSWER")
        if validity_status(item) == "EXPIRED":
            warnings.append("EXPIRED_TEMPORARY")
        if not item.get("active"):
            warnings.append("INACTIVE")

        rows.append({
            "Learning ID": str(item["id"]),
            "상태": "활성" if item.get("active") else "비활성",
            "연결 문의 ID": str(item.get("inquiry_external_id") or ""),
            "질문": mask(item.get("question_original_masked")),
            "학습 답변": mask(answer),
            "상품명": item.get("product_name") or "",
            "상품 ID": product_id,
            "모델 코드": model_code,
            "적용 범위": metadata.get("product_scope") or "",
            "Positive/Negative": signal,
            "문체 전용(style_only)": yes_no(style_only),
            "직원 검증(human_verified)": yes_no(human_verified),
            "직원 승인": yes_no(approved),
            "직원 수정": yes_no(edited),
            "사실 근거 사용 가능": yes_no(fact_eligible),
            "출처(learning_source)": source,
            "근거 등급(facts_authority)": metadata.get("facts_authority") or "",
            "답변 출처(provenance)": metadata.get("answer_provenance") or "",
            "평점": item.get("rating"),
            "품질 점수": round(float(item.get("quality_score") or 0), 3),
            "기간성": yes_no(
                str(item.get("validity_type") or "PERMANENT").upper() == "TEMPORARY"
            ),
            "이벤트명": item.get("event_name") or "",
            "유효 시작": korean_datetime(item.get("valid_from")),
            "유효 종료": korean_datetime(item.get("valid_until")),
            "유효 상태": validity_summary(item),
            "운영 메모": mask(item.get("validity_note")),
            "사용 횟수": int(item.get("usage_count") or 0),
            "마지막 사용": korean_datetime(item.get("last_used_at")),
            "생성일": korean_datetime(item.get("created_at")),
            "수정일": korean_datetime(item.get("updated_at")),
            "데이터 품질 경고": ", ".join(warnings),
            "_temporary": str(item.get("validity_type") or "").upper() == "TEMPORARY",
            "_hedge_reason": hedge_reason(answer) or "",
        })
    return rows


# --------------------------------------------------------------------------
# workbook
# --------------------------------------------------------------------------


COLUMN_WIDTHS = {
    "문의 내용": 60, "프로그램 생성 답변": 60, "직원 수정 답변": 60,
    "최종 답변": 60, "실제 등록 답변": 60, "질문": 45, "학습 답변": 60,
    "상품명": 34, "운영 메모": 30, "보류 사유": 40, "차단 사유 코드": 34,
    "데이터 품질 경고": 30, "사용 Rule/Template": 26, "사용 Learning ID": 20,
}
WRAP_COLUMNS = frozenset({
    "문의 내용", "프로그램 생성 답변", "직원 수정 답변", "최종 답변",
    "실제 등록 답변", "질문", "학습 답변", "운영 메모", "보류 사유",
})


def write_table(
    workbook: Workbook, title: str, headers: list[str], rows: Iterable[dict[str, Any]]
) -> int:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30

    count = 0
    for row in rows:
        count += 1
        values = []
        for header in headers:
            value = row.get(header, "")
            if header in TEXT_COLUMNS:
                value = str(value or "")
            values.append(value)
        sheet.append(values)
        excel_row = count + 1
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=excel_row, column=index)
            cell.border = CELL_BORDER
            if header in TEXT_COLUMNS:
                # Keep identifiers as text so Excel never renders an order
                # number as 2.02608E+15.
                cell.number_format = "@"
            cell.alignment = Alignment(
                vertical="top", wrap_text=header in WRAP_COLUMNS
            )

    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(
            header, max(10, min(len(header) + 6, 24))
        )
    sheet.freeze_panes = "A2"
    if count:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{count + 1}"
        )
    return count


def write_summary(
    workbook: Workbook,
    *,
    qa_rows: list[dict[str, Any]],
    learning_rows: list[dict[str, Any]],
    sources: dict[str, str],
) -> None:
    sheet = workbook.create_sheet("요약", 0)
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 46

    def section(label: str) -> None:
        sheet.append([])
        sheet.append([label])
        sheet.cell(row=sheet.max_row, column=1).font = TITLE_FONT

    def line(label: str, value: object) -> None:
        sheet.append([label, value])

    sheet.append(["Q&A Auto 운영 데이터 Export"])
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    line("생성 시각", datetime.now().strftime("%Y-%m-%d %H:%M"))
    for label, value in sources.items():
        line(label, value)

    section("Q&A")
    line("문의 총 건수", len(qa_rows))
    line("직원검토 필요", sum(1 for row in qa_rows if row["_staff_review"]))
    line("자동등록 가능(등록 대기)",
         sum(1 for row in qa_rows if row["자동등록 판정"] == "자동등록 가능"))
    line("자동등록으로 등록됨",
         sum(1 for row in qa_rows if row["자동등록 여부"] == "예"))
    line("등록 실패", sum(1 for row in qa_rows
                       if str(row["등록 상태"]).upper() == "POST_FAILED"))
    line("답변 초안 없음",
         sum(1 for row in qa_rows if "NO_DRAFT" in row["데이터 품질 경고"]))
    line("저장된 답변에 내부 표시 포함",
         sum(1 for row in qa_rows if "INTERNAL_PLACEHOLDER" in row["데이터 품질 경고"]))

    section("Learning")
    line("Learning 총 건수", len(learning_rows))
    line("활성", sum(1 for row in learning_rows if row["상태"] == "활성"))
    line("비활성", sum(1 for row in learning_rows if row["상태"] == "비활성"))
    line("Positive", sum(1 for row in learning_rows
                        if row["Positive/Negative"] == "POSITIVE"))
    line("Positive 외", sum(1 for row in learning_rows
                          if row["Positive/Negative"] != "POSITIVE"))
    line("문체 전용(style_only)",
         sum(1 for row in learning_rows if row["문체 전용(style_only)"] == "예"))
    line("직원 검증(human_verified)",
         sum(1 for row in learning_rows if row["직원 검증(human_verified)"] == "예"))
    line("사실 근거 사용 가능",
         sum(1 for row in learning_rows if row["사실 근거 사용 가능"] == "예"))
    line("기간성 Learning", sum(1 for row in learning_rows if row["_temporary"]))

    section("데이터 품질 (Learning)")
    for code in ("INTERNAL_PLACEHOLDER", "HEDGED_ANSWER", "MISSING_MODEL_CODE",
                 "MISSING_PRODUCT_IDENTITY", "EXPIRED_TEMPORARY"):
        line(code, sum(1 for row in learning_rows
                       if code in row["데이터 품질 경고"]))

    section("안내")
    sheet.append([
        "각 항목의 뜻은 [데이터 설명] 시트를 참고하세요.",
    ])
    sheet.append([
        "이 파일은 읽기 전용으로 생성되었으며 운영 데이터는 변경되지 않았습니다.",
    ])
    sheet.freeze_panes = "A2"


GLOSSARY: tuple[tuple[str, str, str], ...] = (
    ("Q&A", "문의 ID", "네이버에서 부여한 문의 번호입니다. 네이버 화면에서 같은 번호로 찾을 수 있습니다."),
    ("Q&A", "문의 분류 / 세부 분류", "고객이 무엇을 묻는지 프로그램이 판단한 종류입니다. 예: 배송 일정 문의, 상품 사양 문의."),
    ("Q&A", "Intent", "문의의 핵심 의도입니다. 예: DELIVERY_DATE(배송일), INSTALLATION_DATE(설치일)."),
    ("Q&A", "답변 전략", "그 문의를 어떻게 답할지 정한 방식입니다. 예: REQUEST_ORDER_ID(주문번호 요청)."),
    ("Q&A", "주문조회 필요", "고객 개인의 주문 정보를 네이버에서 조회해야 답할 수 있는 문의인지 여부입니다."),
    ("Q&A", "DPS 조회 필요", "실제 배송·설치 일정을 DPS에서 확인해야 하는지 여부입니다. 주문번호가 없으면 조회하지 않습니다."),
    ("Q&A", "DPS 결과", "DPS 조회 결과입니다. SUCCESS면 조회 성공, 그 외는 조회하지 못한 사유입니다."),
    ("Q&A", "답변 출처", "답변을 만든 방식입니다. RULE=확정 운영 규칙, TEMPLATE=고정 서식, GPT_*=AI 생성."),
    ("Q&A", "사용 Rule/Template", "그 답변을 만든 규칙 또는 서식의 이름입니다."),
    ("Q&A", "Product Fact 사용", "상품 정보 DB에서 검증된 사양을 근거로 사용했는지 여부입니다."),
    ("Q&A", "Learning 사용", "과거 답변 사례를 실제로 근거로 사용했는지 여부입니다."),
    ("Q&A", "프로그램 생성 답변", "프로그램이 처음 만든 답변입니다."),
    ("Q&A", "직원 수정 답변", "직원이 고쳐 쓴 답변입니다. 비어 있으면 수정하지 않은 것입니다."),
    ("Q&A", "최종 답변", "실제로 등록하기로 확정된 답변입니다."),
    ("Q&A", "실제 등록 답변", "네이버에 실제로 올라가 있는 답변 본문입니다."),
    ("Q&A", "Validator", "만들어진 답변에 문제가 없는지 검사한 결과입니다. 통과해야 등록할 수 있습니다."),
    ("Q&A", "자동등록 판정", "이 답변을 사람 없이 등록해도 되는지에 대한 최종 판단입니다."),
    ("Q&A", "차단 사유 코드", "자동등록을 막은 이유를 시스템 코드로 적은 것입니다."),
    ("Q&A", "보류 사유", "위 코드를 사람이 읽을 수 있게 풀어 쓴 문장입니다."),
    ("Q&A", "직원검토 필요", "예이면 직원이 직접 확인해야 하는 문의입니다."),
    ("Q&A", "자동등록 여부", "프로그램이 자동으로 등록했는지 여부입니다. 아니오면 직원이 등록했거나 아직 등록 전입니다."),
    ("Learning", "Learning", "과거 답변 사례를 모아 두었다가 비슷한 새 문의에서 참고하는 데이터입니다. GPT를 재학습시키는 것이 아닙니다."),
    ("Learning", "Positive / Negative", "Positive는 앞으로 참고할 좋은 사례, Negative는 다시 쓰지 않도록 표시한 사례입니다."),
    ("Learning", "문체 전용(style_only)", "예이면 말투와 형식만 참고합니다. 사실 근거로는 쓸 수 없습니다."),
    ("Learning", "직원 검증(human_verified)", "직원이 내용을 확인하고 승인한 사례인지 여부입니다."),
    ("Learning", "직원 승인 / 직원 수정", "승인은 직원이 확인했다는 뜻이고, 수정은 문구를 고쳤다는 뜻입니다. 수정하지 않고 승인해도 정상 승인입니다."),
    ("Learning", "사실 근거 사용 가능", "예이면 이 사례를 새 답변의 근거로 쓸 수 있습니다. 활성·직원 검증·문체 전용 아님·오염 없음·추정 표현 없음·상품 식별 가능을 모두 만족해야 합니다."),
    ("Learning", "적용 범위", "이 사례가 어디까지 적용되는지입니다. MODEL/VARIANT는 특정 모델, POLICY는 상품과 무관한 회사 정책입니다."),
    ("Learning", "기간성", "예이면 정해진 기간에만 유효한 사례입니다. 행사·이벤트 안내 등이 해당합니다."),
    ("Learning", "유효 상태", "기간성 사례의 현재 상태입니다. 활성/시작 전/만료/수동 비활성으로 표시됩니다."),
    ("Learning", "사용 횟수 / 마지막 사용", "이 사례가 실제로 답변에 참고된 횟수와 가장 최근 시점입니다."),
    ("공통", "Provenance", "이 데이터가 어디에서 만들어졌는지를 나타내는 출처 정보입니다."),
    ("공통", "INTERNAL_PLACEHOLDER", "저장된 답변 자체에 <masked-phone> 같은 내부 표시가 들어 있다는 뜻입니다. 고객에게 보이면 안 되는 값이므로 정리가 필요합니다."),
    ("공통", "화면의 <masked-...> 표시", "셀에 <masked-phone> 등이 보여도 두 가지 경우가 있습니다. (1) 이 파일을 만들 때 고객 전화번호를 가린 것 — 정상입니다. (2) 저장된 데이터에 원래 들어 있던 것 — 이 경우에만 '데이터 품질 경고'에 INTERNAL_PLACEHOLDER가 표시됩니다. 경고가 없으면 정리할 필요가 없습니다."),
    ("공통", "HEDGED_ANSWER", "'가능할 것으로 보입니다'처럼 확정하지 않은 표현입니다. 사실 근거로는 사용하지 않습니다."),
    ("공통", "MISSING_MODEL_CODE", "모델 코드가 비어 있어 같은 모델인지 확인하기 어렵습니다."),
    ("공통", "MISSING_PRODUCT_IDENTITY", "어떤 상품의 사례인지 식별할 정보가 없습니다."),
    ("공통", "EXPIRED_TEMPORARY", "기간이 지난 기간성 사례입니다. 검색에서는 제외되지만 기록은 남아 있습니다."),
    ("공통", "개인정보 표시", "고객 이름·전화번호·주소·이메일은 저장 시점과 이 파일 생성 시점에 모두 가려집니다. 1588-3366, 02-706-2678 같은 공식 고객센터 번호는 그대로 보입니다."),
)


def write_glossary(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("데이터 설명")
    headers = ["구분", "항목", "설명"]
    sheet.append(headers)
    for index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")
    for group, term, description in GLOSSARY:
        sheet.append([group, term, description])
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=3).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        sheet.cell(row=row, column=1).alignment = Alignment(vertical="top")
        sheet.cell(row=row, column=2).alignment = Alignment(vertical="top")
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 96
    sheet.freeze_panes = "A2"


TEMPORARY_HEADERS = [
    "Learning ID", "상태", "이벤트명", "질문", "학습 답변", "상품명", "상품 ID",
    "적용 범위", "유효 시작", "유효 종료", "유효 상태", "운영 메모",
    "생성일", "데이터 품질 경고",
]

REVIEW_HEADERS = [
    "문의 ID", "문의 등록일", "상품명", "문의 내용",
    "문의 분류", "Intent", "답변 출처", "사용 Rule/Template",
    "프로그램 생성 답변", "직원 수정 답변", "최종 답변",
    "Validator", "자동등록 판정", "보류 사유", "차단 사유 코드",
    "DPS 결과", "데이터 품질 경고",
]


def build_workbook(
    *,
    qa_rows: list[dict[str, Any]],
    learning_rows: list[dict[str, Any]],
    sources: dict[str, str],
) -> tuple[Workbook, dict[str, int]]:
    workbook = Workbook()
    workbook.remove(workbook.active)

    counts: dict[str, int] = {}
    counts["Q&A 전체"] = write_table(workbook, "Q&A 전체", QA_HEADERS, qa_rows)
    counts["Learning 전체"] = write_table(
        workbook, "Learning 전체", LEARNING_HEADERS, learning_rows
    )
    counts["직원검토"] = write_table(
        workbook, "직원검토", REVIEW_HEADERS,
        [row for row in qa_rows if row["_staff_review"]],
    )
    counts["기간성 Learning"] = write_table(
        workbook, "기간성 Learning", TEMPORARY_HEADERS,
        [row for row in learning_rows if row["_temporary"]],
    )
    write_glossary(workbook)
    write_summary(
        workbook, qa_rows=qa_rows, learning_rows=learning_rows, sources=sources
    )
    workbook._sheets.sort(
        key=lambda sheet: [
            "요약", "Q&A 전체", "Learning 전체", "직원검토",
            "기간성 Learning", "데이터 설명",
        ].index(sheet.title)
    )
    return workbook, counts


def export(
    *, out_dir: Path | None = None, limit: int | None = None
) -> tuple[Path, dict[str, int]]:
    """Build the workbook and return its path plus the per-sheet row counts."""

    db_path = Path(get_database_path()).resolve()
    facts_path = Path(get_product_facts_path()).resolve()
    before = {"qa": fingerprint(db_path), "facts": fingerprint(facts_path)}

    connection = open_readonly(db_path)
    try:
        qa_rows = load_qa_rows(connection, limit=limit)
        learning_rows = load_learning_rows(connection)
    finally:
        connection.close()

    sources = {
        "운영 데이터베이스": str(db_path),
        "Q&A 건수(전체 조회)": str(len(qa_rows)),
        "Learning 건수(전체 조회)": str(len(learning_rows)),
    }
    workbook, counts = build_workbook(
        qa_rows=qa_rows, learning_rows=learning_rows, sources=sources
    )

    target_dir = Path(out_dir) if out_dir else PROJECT_ROOT / "exports"
    target_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = target_dir / f"Q&A_Auto_운영데이터_{stamp}.xlsx"
    workbook.save(path)

    after = {"qa": fingerprint(db_path), "facts": fingerprint(facts_path)}
    if before != after:
        raise RuntimeError(
            "DB_MODIFIED_DURING_EXPORT: 운영 데이터가 변경되었습니다. "
            f"before={before} after={after}"
        )
    return path, counts


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out-dir", default=None)
    parser.add_argument("--limit", type=int, default=None,
                        help="Q&A 시트 최대 행 수 (기본: 전체)")
    arguments = parser.parse_args()

    path, counts = export(
        out_dir=Path(arguments.out_dir) if arguments.out_dir else None,
        limit=arguments.limit,
    )
    print("Export 완료")
    print(f"  파일: {path}")
    for sheet, count in counts.items():
        print(f"  {sheet}: {count}행")
    print("  운영 DB 변경 없음 (크기·수정시각 동일)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
