from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from repositories.database import Database
from repositories.learning_repository import LearningRepository
from services.learning_privacy_service import LearningPrivacyService
from services.learning_quality_service import LearningQualityService
from services.learning_service import LearningService
from services.similar_answer_service import normalize_learning_question


ANSWER_HISTORY = Path("outputs") / "learning" / "answer_history.jsonl"
NAVER_OUTPUTS = Path("outputs") / "naver_api"
PLACEHOLDER_ANSWERS = {
    "", "자동답변 안함", "NOT_SUPPORTED", "ERROR", "답변하지 않음",
}
POSTED_VALUES = {"1", "TRUE", "Y", "YES"}
LEGACY_SOURCES = {"GPT": "LEGACY_GPT", "RULE": "LEGACY_RULE", "RULES": "LEGACY_RULE"}


def _text(value: object) -> str:
    return str(value or "").strip()


def _posted(value: object) -> bool:
    return _text(value).upper() in POSTED_VALUES


def _parse_raw(value: object) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def _customer_names(value: Any) -> set[str]:
    found: set[str] = set()
    if isinstance(value, dict):
        for key, item in value.items():
            normalized = str(key).replace("-", "").replace("_", "").lower()
            if normalized in {
                "customername", "membername", "username", "writername",
                "customerdisplay", "maskedwriterid",
            } and isinstance(item, str) and 2 <= len(item.strip()) <= 30:
                found.add(item.strip())
            found.update(_customer_names(item))
    elif isinstance(value, list):
        for item in value:
            found.update(_customer_names(item))
    return found


def _iter_legacy_rows(legacy_root: Path) -> Iterable[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("Legacy XLSX 조사에는 openpyxl이 필요합니다.") from exc
    for workbook_path in sorted((legacy_root / NAVER_OUTPUTS).glob("*.xlsx")):
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            worksheet = workbook["naver_qna_result"]
            rows = worksheet.iter_rows(values_only=True)
            headers = [_text(value) for value in next(rows, ())]
            for values in rows:
                row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
                row["_source_file"] = workbook_path.name
                yield row
        finally:
            workbook.close()


def inspect_answer_history(legacy_root: Path) -> dict[str, int]:
    usable = excluded = invalid = 0
    path = legacy_root / ANSWER_HISTORY
    if not path.exists():
        return {"usable": 0, "excluded": 0, "invalid": 0}
    with path.open("r", encoding="utf-8") as stream:
        for line in stream:
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                invalid += 1
                continue
            answer = _text(row.get("program_answer"))
            eligible = (
                bool(_text(row.get("question")))
                and answer.upper() not in PLACEHOLDER_ANSWERS
                and _text(row.get("action")).lower() == "posted"
                and _posted(row.get("posted"))
                and _text(row.get("program_status")) == "답변 가능"
            )
            usable += int(eligible)
            excluded += int(not eligible)
    return {"usable": usable, "excluded": excluded, "invalid": invalid}


def build_legacy_candidates(legacy_root: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    privacy, quality = LearningPrivacyService(), LearningQualityService()
    candidates: list[dict[str, Any]] = []
    exclusions: Counter[str] = Counter()
    scanned = answer_present = gpt_answers = rule_answers = 0
    seen: set[str] = set()
    files: set[str] = set()
    for row in _iter_legacy_rows(legacy_root):
        scanned += 1
        files.add(_text(row.get("_source_file")))
        question, answer = _text(row.get("문의")), _text(row.get("답변"))
        provider = _text(row.get("판단방식")).upper()
        if answer:
            answer_present += 1
            gpt_answers += int(provider == "GPT")
            rule_answers += int(provider in {"RULE", "RULES"})
        reason = ""
        if not question:
            reason = "EMPTY_QUESTION"
        elif not answer or answer.upper() in PLACEHOLDER_ANSWERS:
            reason = "EMPTY_OR_UNSUPPORTED_ANSWER"
        elif _text(row.get("답변여부")) != "답변 가능":
            reason = "VALIDATOR_FAILED"
        elif _text(row.get("action")).lower() != "posted" or not _posted(row.get("posted")):
            reason = "POSTED_FAILED"
        elif provider not in LEGACY_SOURCES:
            reason = "UNSUPPORTED_PROVIDER"
        if reason:
            exclusions[reason] += 1
            continue
        raw = _parse_raw(row.get("raw_json"))
        names = _customer_names(raw)
        masked_question = privacy.mask(question, customer_names=names)
        masked_answer = privacy.mask(answer, customer_names=names)
        legacy_source = LEGACY_SOURCES[provider]
        source_key = hashlib.sha256(
            f"{legacy_source}|{masked_question}|{masked_answer}".encode("utf-8")
        ).hexdigest()
        if source_key in seen:
            exclusions["DUPLICATE"] += 1
            continue
        seen.add(source_key)
        rating = 4 if legacy_source == "LEGACY_RULE" else 3
        created = _text(row.get("처리시각"))
        candidates.append({
            "source_key": source_key,
            "inquiry_id": None,
            "answer_draft_id": None,
            "approval_history_id": None,
            "learning_source": "SELLER_ANSWER",
            "question_original_masked": masked_question,
            "question_normalized": normalize_learning_question(masked_question),
            "store_code": "OJE_PLUS",
            "inquiry_type": _text(row.get("질문유형")) or None,
            "intent": _text(row.get("질문유형")) or None,
            "product_name": privacy.mask(row.get("상품")) or None,
            "model_code": privacy.mask(row.get("옵션")) or None,
            "generation_mode": legacy_source,
            "template_id": None,
            "processing_route": legacy_source,
            "validator_result": "LEGACY_POSTED_PASS",
            "seller_answer": masked_answer,
            "gpt_draft": masked_answer if legacy_source == "LEGACY_GPT" else None,
            "edited_answer": None,
            "final_answer": masked_answer,
            "posted": True,
            "posted_at": created or None,
            "auto_posted": False,
            "rating": rating,
            "edit_ratio": 0.0,
            "quality_score": rating / 5,
            "style_only": True,
            "version": 1,
            "style_features_json": quality.style_features(masked_answer),
            "metadata_json": {
                "facts_authority": "STYLE_ONLY",
                "legacy_source": legacy_source,
                "source_file": _text(row.get("_source_file")),
                "source_question_id_hash": hashlib.sha256(
                    _text(row.get("question_id")).encode("utf-8")
                ).hexdigest(),
                "raw_json_present": bool(raw),
            },
            "active": True,
        })
    summary = {
        "files": len(files), "inquiries": scanned,
        "answer_present": answer_present, "gpt_answers": gpt_answers,
        "rule_answers": rule_answers, "eligible": len(candidates),
        "excluded": sum(exclusions.values()), "exclusion_reasons": dict(exclusions),
    }
    return candidates, summary


def backfill_legacy(database: Database, legacy_root: Path, *, apply: bool) -> dict[str, Any]:
    repository = LearningRepository(database)
    candidates, legacy_summary = build_legacy_candidates(legacy_root)
    existing = added = duplicates = 0
    for candidate in candidates:
        if repository.get_by_source_key(candidate["source_key"]):
            duplicates += 1
            continue
        if apply:
            repository.upsert(candidate)
            added += 1
    existing = repository.count()
    return {
        "mode": "APPLY" if apply else "DRY_RUN",
        "legacy": legacy_summary,
        "answer_history": inspect_answer_history(legacy_root),
        "repository_existing": existing - added,
        "repository_added": added if apply else len(candidates) - duplicates,
        "repository_final": existing,
        "duplicates": duplicates,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="승인 답변과 동기화된 판매자 답변을 Learning Repository로 가져옵니다.")
    parser.add_argument("--database")
    parser.add_argument("--apply", action="store_true", help="생성/갱신을 실제 적용합니다. 생략하면 건수만 확인합니다.")
    parser.add_argument("--seller-limit", type=int)
    parser.add_argument("--legacy-root", type=Path, help="읽기 전용 Legacy QnA 배포용 프로젝트 경로")
    args = parser.parse_args()
    database = Database(args.database)
    database.initialize()
    if args.legacy_root:
        summary = backfill_legacy(database, args.legacy_root.resolve(), apply=args.apply)
        print(json.dumps(summary, ensure_ascii=False, sort_keys=True))
        return 0
    if not args.apply:
        with database.connection() as connection:
            summary = {
                "mode": "DRY_RUN",
                "approved_candidates": connection.execute(
                    "SELECT COUNT(*) FROM answer_drafts WHERE review_status='APPROVED' AND trim(COALESCE(final_answer,''))<>''"
                ).fetchone()[0],
                "seller_answered_candidates": connection.execute(
                    "SELECT COUNT(*) FROM inquiries WHERE source_answered=1"
                ).fetchone()[0],
            }
    else:
        service = LearningService(database)
        summary = {
            "mode": "APPLY",
            "approved": service.import_existing_approved(),
            "seller": service.import_existing_seller_answers(limit=args.seller_limit),
            "repository_count": service.repository.count(),
        }
    print(json.dumps(summary, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
