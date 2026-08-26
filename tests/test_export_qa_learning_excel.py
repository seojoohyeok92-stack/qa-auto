"""The operational Excel export, checked against a real-shaped database.

Two properties matter more than the contents of any single cell, and both are
asserted here rather than described:

* the export never writes to the operational data -- the fixture database is
  fingerprinted before and after every run;
* it never publishes something the pipeline would have masked -- the same
  ``LearningPrivacyService`` the rest of the system uses is applied on the way
  out, and the approved contact numbers stay readable because that policy says
  so, not because this script decided to.

The fixture is built with the production schema (``Database`` runs the real
migrations) and filled with rows shaped like the operational ones, including
the awkward cases: a sixteen-digit order number Excel would otherwise render
as 2.02608E+15, a stored redaction token, a hedged approved answer, an expired
temporary Learning, and a long Korean body.
"""
from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from repositories.database import Database


ORDER_NUMBER = "2026082351391541"
PRODUCT_ID = "13239109816"
LONG_BODY = (
    "안녕하세요 혹시 배송 올때 조립에 필요한 일회용 공구도 같이 오나요? "
    "제가 개별로 준비해야 하는 공구가 있는지도 알려주세요. " * 6
)
OFFICIAL_ANSWER = "자세한 사항은 삼성전자 고객센터 1588-3366으로 문의해 주세요."
PRIVATE_ANSWER = "담당자 휴대폰 010-1234-5678로 연락 주시면 안내드리겠습니다."
CONTAMINATED_ANSWER = "문의는 <masked-phone>로 연락 주세요."
HEDGED_ANSWER = "해당 기능은 사용 가능할 것으로 보입니다."
DEFINITE_ANSWER = "기본 스탠드는 탈부착 가능합니다."


@pytest.fixture()
def operational_db(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """A database with the production schema and representative rows."""

    path = tmp_path / "oje_automation.db"
    database = Database(path)
    database.initialize()

    with database.connection() as connection:
        connection.execute(
            """
            INSERT INTO inquiries
                (id, store_code, source_type, source_question_id,
                 external_inquiry_id, inquiry_type, title, content,
                 product_name, product_id, order_id, registered_at,
                 workflow_status, answer_status, post_status, raw_json,
                 approval_status)
            VALUES (1,'OJE_PLUS','PRODUCT_INQUIRY','686472270','686472270',
                    'PRODUCT_INQUIRY','상품 문의',?,'삼성 스마트 모니터 M5 32인치',
                    ?,?, '2026-08-26T10:00:00+09:00','REVIEW_PENDING',
                    'UNANSWERED','NOT_POSTED','{}','PENDING')
            """,
            (LONG_BODY, PRODUCT_ID, ORDER_NUMBER),
        )
        connection.execute(
            """
            INSERT INTO inquiries
                (id, store_code, source_type, source_question_id,
                 external_inquiry_id, inquiry_type, title, content,
                 product_name, registered_at, workflow_status, answer_status,
                 post_status, raw_json, approval_status, source_answered)
            VALUES (2,'OJE_PLUS','CUSTOMER_INQUIRY','325262026','325262026',
                    'CUSTOMER_INQUIRY','고객 문의','배송 언제 되나요?',
                    '삼성 43인치 TV','2026-08-26T11:00:00+09:00','POSTED',
                    'ANSWERED','POSTED','{}','APPROVED',1)
            """
        )
        connection.execute(
            """
            INSERT INTO answer_drafts
                (id, inquiry_id, program_status, category, reason, provider,
                 original_answer, final_answer, review_status, posted,
                 validation_status, is_active, metadata_json)
            VALUES (1,1,'NEEDS_REVIEW','PRODUCT_GENERAL','테스트','phase9_policy',
                    ?,?, 'NEEDS_REVIEW',0,'REVIEW_REQUIRED',1,
                    '{"generation_mode":"RULE","template_id":"T1",
                      "processing_plan":{"analysis":{"question_category":"PRODUCT_GENERAL",
                      "detected_intent":"GENERAL","requires_order_lookup":false,
                      "requires_dps_lookup":false},"order_id_status":"VALID",
                      "needs_staff_review":true}}')
            """,
            (CONTAMINATED_ANSWER, CONTAMINATED_ANSWER),
        )
        connection.execute(
            """
            INSERT INTO naver_posted_answers
                (id, source_key, inquiry_id, answer_body, fetch_status,
                 posted_at, is_current, source_api,
                 first_observed_at, last_observed_at)
            VALUES (1,'k1',2,?,'AVAILABLE','2026-08-26T11:30:00+09:00',1,
                    'CUSTOMER_INQUIRY','2026-08-26T11:30:00Z',
                    '2026-08-26T11:30:00Z')
            """,
            (PRIVATE_ANSWER,),
        )
        for index, (answer, style_only, verified, validity) in enumerate(
            [
                (DEFINITE_ANSWER, 0, 1, "PERMANENT"),
                (HEDGED_ANSWER, 0, 1, "PERMANENT"),
                (CONTAMINATED_ANSWER, 0, 1, "PERMANENT"),
                (OFFICIAL_ANSWER, 1, 0, "PERMANENT"),
                (DEFINITE_ANSWER, 0, 1, "TEMPORARY"),
            ],
            start=1,
        ):
            metadata = (
                '{"human_verified": true, "facts_authority": "APPROVED_REFERENCE",'
                ' "learning_signal_type": "POSITIVE",'
                f' "product_identity": {{"product_id": "{PRODUCT_ID}"}}}}'
                if verified
                else '{"facts_authority": "STYLE_ONLY",'
                     ' "learning_signal_type": "POSITIVE"}'
            )
            connection.execute(
                """
                INSERT INTO learning_examples
                    (id, source_key, inquiry_id, learning_source,
                     question_original_masked, question_normalized, store_code,
                     inquiry_type, product_name, final_answer, rating,
                     style_only, metadata_json, active, validity_type,
                     event_name, valid_from, valid_until, created_at)
                VALUES (?,?,1,'APPROVED_UNEDITED','질문입니다','질문입니다','OJE_PLUS',
                        'PRODUCT_INQUIRY','삼성 스마트 모니터 M5 32인치',?,5,?,?,1,?,
                        ?,?,?, '2026-08-20T09:00:00Z')
                """,
                (
                    index, f"key-{index}", answer, style_only, metadata, validity,
                    "삼성 감사제" if validity == "TEMPORARY" else None,
                    "2026-07-01" if validity == "TEMPORARY" else None,
                    "2026-07-31" if validity == "TEMPORARY" else None,
                ),
            )
        connection.commit()

    monkeypatch.setenv("OJE_AUTOMATION_DB_PATH", str(path))
    return path


def fingerprint(path: Path) -> tuple[int, int]:
    stat = path.stat()
    return stat.st_size, stat.st_mtime_ns


def run_export(tmp_path: Path) -> tuple[Path, dict[str, int]]:
    from scripts import export_qa_learning_excel as exporter

    return exporter.export(out_dir=tmp_path / "exports")


# --------------------------------------------------------------------------


def test_ex01_export_succeeds_against_a_real_schema_database(
    operational_db: Path, tmp_path: Path
) -> None:
    path, counts = run_export(tmp_path)

    assert path.exists()
    assert path.stat().st_size > 0
    assert counts["Q&A 전체"] == 2


def test_ex02_workbook_reopens_cleanly(
    operational_db: Path, tmp_path: Path
) -> None:
    path, _ = run_export(tmp_path)

    workbook = load_workbook(path)

    assert workbook.sheetnames


def test_ex03_required_sheets_exist_in_order(
    operational_db: Path, tmp_path: Path
) -> None:
    path, _ = run_export(tmp_path)

    assert load_workbook(path).sheetnames == [
        "요약", "Q&A 전체", "Learning 전체", "직원검토",
        "기간성 Learning", "데이터 설명",
    ]


def test_ex04_qa_row_count_matches_the_database(
    operational_db: Path, tmp_path: Path
) -> None:
    path, counts = run_export(tmp_path)
    with sqlite3.connect(operational_db) as connection:
        expected = connection.execute("SELECT COUNT(*) FROM inquiries").fetchone()[0]

    assert counts["Q&A 전체"] == expected
    assert load_workbook(path)["Q&A 전체"].max_row == expected + 1


def test_ex05_learning_row_count_matches_the_database(
    operational_db: Path, tmp_path: Path
) -> None:
    path, counts = run_export(tmp_path)
    with sqlite3.connect(operational_db) as connection:
        expected = connection.execute(
            "SELECT COUNT(*) FROM learning_examples"
        ).fetchone()[0]

    assert counts["Learning 전체"] == expected
    assert load_workbook(path)["Learning 전체"].max_row == expected + 1


def test_ex06_long_korean_text_survives_intact(
    operational_db: Path, tmp_path: Path
) -> None:
    path, _ = run_export(tmp_path)
    sheet = load_workbook(path)["Q&A 전체"]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index("문의 내용") + 1
    bodies = [
        sheet.cell(row=row, column=column).value or ""
        for row in range(2, sheet.max_row + 1)
    ]

    assert any(len(body) > 300 for body in bodies)
    assert any("조립에 필요한 일회용 공구" in body for body in bodies)


def test_ex07_order_number_is_not_turned_into_scientific_notation(
    operational_db: Path, tmp_path: Path
) -> None:
    """2026082351391541 must stay searchable, not become 2.02608E+15."""

    path, _ = run_export(tmp_path)
    sheet = load_workbook(path)["Q&A 전체"]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index("주문번호") + 1
    values = [
        sheet.cell(row=row, column=column).value
        for row in range(2, sheet.max_row + 1)
    ]

    assert ORDER_NUMBER in values
    stored = sheet.cell(row=values.index(ORDER_NUMBER) + 2, column=column)
    assert isinstance(stored.value, str)
    assert stored.number_format == "@"


def test_ex08_dates_use_one_readable_format(
    operational_db: Path, tmp_path: Path
) -> None:
    path, _ = run_export(tmp_path)
    sheet = load_workbook(path)["Q&A 전체"]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index("문의 등록일") + 1
    values = [
        sheet.cell(row=row, column=column).value
        for row in range(2, sheet.max_row + 1)
    ]

    assert values
    for value in values:
        assert value == "" or len(value) == len("2026-08-26 10:00"), value


def test_ex09_private_contact_details_are_masked(
    operational_db: Path, tmp_path: Path
) -> None:
    """A live mobile number stored in a posted answer must not reach Excel."""

    path, _ = run_export(tmp_path)
    workbook = load_workbook(path)
    text = "\n".join(
        str(value)
        for name in workbook.sheetnames
        for row in workbook[name].iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    )

    assert "010-1234-5678" not in text
    assert "<masked-phone>" in text


def test_ex10_approved_contact_numbers_stay_readable(
    operational_db: Path, tmp_path: Path
) -> None:
    path, _ = run_export(tmp_path)
    workbook = load_workbook(path)
    text = "\n".join(
        str(value)
        for name in workbook.sheetnames
        for row in workbook[name].iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    )

    assert "1588-3366" in text


def test_ex11_stored_redaction_tokens_are_flagged(
    operational_db: Path, tmp_path: Path
) -> None:
    """A token already in the database is a finding; export masking is not."""

    path, _ = run_export(tmp_path)
    workbook = load_workbook(path)

    learning = workbook["Learning 전체"]
    headers = [cell.value for cell in learning[1]]
    answer_column = headers.index("학습 답변") + 1
    warning_column = headers.index("데이터 품질 경고") + 1
    flagged = 0
    for row in range(2, learning.max_row + 1):
        answer = learning.cell(row=row, column=answer_column).value or ""
        warning = learning.cell(row=row, column=warning_column).value or ""
        if "<masked-" in answer:
            assert "INTERNAL_PLACEHOLDER" in warning
            flagged += 1
    assert flagged == 1

    # The posted answer masked *by this export* carries no such finding.
    qa = workbook["Q&A 전체"]
    qa_headers = [cell.value for cell in qa[1]]
    posted_column = qa_headers.index("실제 등록 답변") + 1
    qa_warning = qa_headers.index("데이터 품질 경고") + 1
    for row in range(2, qa.max_row + 1):
        if "<masked-phone>" in (qa.cell(row=row, column=posted_column).value or ""):
            assert "INTERNAL_PLACEHOLDER" not in (
                qa.cell(row=row, column=qa_warning).value or ""
            )


def test_ex11b_other_quality_findings_are_reported(
    operational_db: Path, tmp_path: Path
) -> None:
    path, _ = run_export(tmp_path)
    sheet = load_workbook(path)["Learning 전체"]
    headers = [cell.value for cell in sheet[1]]
    column = headers.index("데이터 품질 경고") + 1
    warnings = "\n".join(
        sheet.cell(row=row, column=column).value or ""
        for row in range(2, sheet.max_row + 1)
    )

    assert "HEDGED_ANSWER" in warnings
    assert "STYLE_ONLY" in warnings
    assert "MISSING_MODEL_CODE" in warnings
    assert "EXPIRED_TEMPORARY" in warnings


def test_ex12_temporary_learning_gets_its_own_sheet(
    operational_db: Path, tmp_path: Path
) -> None:
    path, counts = run_export(tmp_path)
    sheet = load_workbook(path)["기간성 Learning"]
    headers = [cell.value for cell in sheet[1]]

    assert counts["기간성 Learning"] == 1
    assert sheet.cell(row=2, column=headers.index("이벤트명") + 1).value == "삼성 감사제"
    assert sheet.cell(row=2, column=headers.index("유효 상태") + 1).value


def test_ex13_staff_review_sheet_matches_the_production_verdict(
    operational_db: Path, tmp_path: Path
) -> None:
    """The sheet is filtered by the dashboard's own staff-review decision."""

    path, counts = run_export(tmp_path)
    workbook = load_workbook(path)
    qa = workbook["Q&A 전체"]
    headers = [cell.value for cell in qa[1]]
    column = headers.index("직원검토 필요") + 1
    expected = sum(
        1
        for row in range(2, qa.max_row + 1)
        if qa.cell(row=row, column=column).value == "예"
    )

    assert counts["직원검토"] == expected
    assert workbook["직원검토"].max_row == expected + 1


def test_ex14_operational_database_is_not_modified(
    operational_db: Path, tmp_path: Path
) -> None:
    before = fingerprint(operational_db)

    run_export(tmp_path)

    assert fingerprint(operational_db) == before


def test_readability_settings_are_applied(
    operational_db: Path, tmp_path: Path
) -> None:
    path, _ = run_export(tmp_path)
    workbook = load_workbook(path)

    for name in ("Q&A 전체", "Learning 전체"):
        sheet = workbook[name]
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref
        assert sheet.cell(row=1, column=1).font.bold
        assert sheet.column_dimensions["A"].width


def test_glossary_explains_every_warning_code(
    operational_db: Path, tmp_path: Path
) -> None:
    """A code an operator can see must be a code the workbook explains."""

    path, _ = run_export(tmp_path)
    workbook = load_workbook(path)
    glossary = "\n".join(
        str(value)
        for row in workbook["데이터 설명"].iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    )

    for code in (
        "INTERNAL_PLACEHOLDER", "HEDGED_ANSWER", "MISSING_MODEL_CODE",
        "MISSING_PRODUCT_IDENTITY", "EXPIRED_TEMPORARY",
    ):
        assert code in glossary, code


def test_filename_is_valid_on_windows(
    operational_db: Path, tmp_path: Path
) -> None:
    path, _ = run_export(tmp_path)

    assert path.suffix == ".xlsx"
    assert path.name.startswith("Q&A_Auto_운영데이터_")
    assert not set(path.name) & set('<>:"/\\|?*')
